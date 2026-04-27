#!/usr/bin/env python3
"""Scrape SITUR (Tourism Ministry) XLSX files → combined CSV.
Processes: listaAgentii, listaCazari, listaGhizi, listaStatiuni, listaCentreInformare.
"""
import csv
import json
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "DATA"
OUT_CSV = DATA_DIR / "situr_combined.csv"

FILES_TO_PROCESS = [
    ("listaAgentii.xlsx", "agentie_turism"),
    ("listaCazari.xlsx", "cazare"),
    ("listaGhizi.xlsx", "ghid_turism"),
    ("listaStatiuni.xlsx", "statiune_turistica"),
    ("listaCentreInformare.xlsx", "centru_informare"),
]


def clean_str(val):
    """Clean and return string value."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def clean_cui(raw):
    """Extract digits from CUI, remove leading zeros."""
    if not raw:
        return None
    import re
    s = re.sub(r"[^\d]", "", str(raw)).lstrip("0")
    return s if s else None


def load_xlsx_rows(path):
    """Load non-empty rows from XLSX, skip first 2 rows (title + header)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(row)
        wb.close()
        return rows[2:] if len(rows) >= 2 else []
    except Exception as e:
        print(f"ERROR loading {path}: {e}")
        return []


def parse_agentii(rows):
    """Parse listaAgentii.xlsx (travel agencies)."""
    records = []
    for r in rows:
        if not r or len(r) < 13:
            continue
        rec = {
            "source_file": "listaAgentii",
            "company_name": clean_str(r[5] or r[3] or ""),
            "cui": clean_cui(r[6]),
            "adresa": clean_str(r[7]),
            "localitate": clean_str(r[8]),
            "judet": clean_str(r[9]),
            "tip": "agentie_turism",
            "email": clean_str(r[12]),
            "extra": json.dumps({
                "operator": clean_str(r[3]),
                "tip": clean_str(r[10])
            }, ensure_ascii=False)
        }
        if rec["company_name"]:
            records.append(rec)
    return records


def parse_cazari(rows):
    """Parse listaCazari.xlsx (accommodations)."""
    records = []
    for r in rows:
        if not r or len(r) < 14:
            continue
        rec = {
            "source_file": "listaCazari",
            "company_name": clean_str(r[1] or ""),
            "cui": clean_cui(r[15] if len(r) > 15 else None),
            "adresa": clean_str(r[5] if len(r) > 5 else None),
            "localitate": clean_str(r[8] if len(r) > 8 else None),
            "judet": clean_str(r[9] if len(r) > 9 else None),
            "tip": "cazare",
            "email": clean_str(r[10] if len(r) > 10 else None),
            "extra": json.dumps({
                "tip_unitate": clean_str(r[0]),
                "categorie": clean_str(r[2]),
                "operator": clean_str(r[13] if len(r) > 13 else None)
            }, ensure_ascii=False)
        }
        if rec["company_name"]:
            records.append(rec)
    return records


def parse_ghizi(rows):
    """Parse listaGhizi.xlsx (tour guides)."""
    records = []
    for r in rows:
        if not r or len(r) < 2:
            continue
        rec = {
            "source_file": "listaGhizi",
            "company_name": clean_str(r[0] or ""),
            "cui": None,
            "adresa": None,
            "localitate": None,
            "judet": None,
            "tip": "ghid_turism",
            "email": None,
            "extra": json.dumps({
                "nr_atestat": clean_str(r[1]),
                "tip": clean_str(r[3] if len(r) > 3 else None)
            }, ensure_ascii=False)
        }
        if rec["company_name"]:
            records.append(rec)
    return records


def parse_statiuni(rows):
    """Parse listaStatiuni.xlsx (mountain/seaside resorts)."""
    records = []
    for r in rows:
        if not r or len(r) < 1:
            continue
        rec = {
            "source_file": "listaStatiuni",
            "company_name": clean_str(r[0] or ""),
            "cui": None,
            "adresa": None,
            "localitate": clean_str(r[0]),
            "judet": clean_str(r[2] if len(r) > 2 else None),
            "tip": "statiune_turistica",
            "email": None,
            "extra": json.dumps({
                "tip": clean_str(r[1] if len(r) > 1 else None)
            }, ensure_ascii=False)
        }
        if rec["company_name"]:
            records.append(rec)
    return records


def parse_centre(rows):
    """Parse listaCentreInformare.xlsx (tourist info centers)."""
    records = []
    for r in rows:
        if not r or len(r) < 2:
            continue
        rec = {
            "source_file": "listaCentreInformare",
            "company_name": clean_str(r[0] or ""),
            "cui": None,
            "adresa": clean_str(r[3] if len(r) > 3 else None),
            "localitate": None,
            "judet": None,
            "tip": "centru_informare",
            "email": clean_str(r[6] if len(r) > 6 else None),
            "extra": json.dumps({
                "administrator": clean_str(r[2] if len(r) > 2 else None),
                "telefon": clean_str(r[4] if len(r) > 4 else None)
            }, ensure_ascii=False)
        }
        if rec["company_name"]:
            records.append(rec)
    return records


PARSERS = [
    ("listaAgentii.xlsx", parse_agentii),
    ("listaCazari.xlsx", parse_cazari),
    ("listaGhizi.xlsx", parse_ghizi),
    ("listaStatiuni.xlsx", parse_statiuni),
    ("listaCentreInformare.xlsx", parse_centre),
]


def main():
    """Process all XLSX files and export to CSV."""
    all_records = []
    print("Processing SITUR XLSX files...")

    for fname, parser in PARSERS:
        fpath = DATA_DIR / fname
        if not fpath.exists():
            print(f"  SKIP {fname} (not found)")
            continue

        rows = load_xlsx_rows(fpath)
        records = parser(rows)
        all_records.extend(records)
        print(f"  {fname:30s}: {len(records):5d} records")

    # Dedup by (source_file, company_name, cui)
    seen = set()
    deduped = []
    for rec in all_records:
        key = (rec["source_file"], rec["company_name"], rec["cui"] or "")
        if key not in seen:
            seen.add(key)
            deduped.append(rec)

    # Write CSV
    if deduped:
        with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "source_file", "company_name", "cui", "adresa",
                "localitate", "judet", "tip", "email", "extra"
            ])
            writer.writeheader()
            writer.writerows(deduped)
        print(f"\n=== EXPORT ===")
        print(f"Total records: {len(deduped)}")
        print(f"Output: {OUT_CSV}")

        # Summary by source
        summary = {}
        for rec in deduped:
            src = rec["source_file"]
            summary[src] = summary.get(src, 0) + 1
        print("\nBy source:")
        for src, cnt in sorted(summary.items()):
            print(f"  {src:30s}: {cnt:5d}")
    else:
        print("No records extracted")


if __name__ == "__main__":
    main()
