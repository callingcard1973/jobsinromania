#!/usr/bin/env python3
"""
Scrape ISC (Construction Inspection) registries from Excel files.
Extracts: diriginti (site managers), RTE (technical managers), laboratoare (testing labs).
Output: CSV files per category or consolidated CSV.
Run: cd D:\\MEMORY\\BUSINESS\\IDEAS\\ISCIR\\ISC && python CODE/scrape_isc.py
"""
import re
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "DATA"
OUTPUT_DIR = DATA_DIR

XLSX_FILES = {
    "diriginti": DATA_DIR / "isc_diriginti.xlsx",
    "rte": DATA_DIR / "isc_rte.xlsx",
    "laboratoare": DATA_DIR / "isc_laboratoare.xlsx",
}

OUTPUT_FILES = {
    "diriginti": OUTPUT_DIR / "isc_diriginti_raw.csv",
    "rte": OUTPUT_DIR / "isc_rte_raw.csv",
    "laboratoare": OUTPUT_DIR / "isc_laboratoare_raw.csv",
    "all": OUTPUT_DIR / "isc_constructii_all.csv",
}


def clean(v) -> str:
    """Convert cell value to clean string."""
    if v is None:
        return ""
    return str(v).strip()


def extract_cui(text: str) -> str:
    """Extract CUI from pattern like 'COMPANY NAME (12345678)'."""
    m = re.search(r"\((\d{4,10})\)", text)
    return m.group(1) if m else ""


def clean_company_name(text: str) -> str:
    """Remove CUI suffix from company name."""
    return re.sub(r"\s*\(\d{4,10}\)\s*$", "", text).strip()


def clean_html(text: str) -> str:
    """Remove HTML tags from text (e.g., <br>, <p>, etc)."""
    return re.sub(r"<[^>]+>", "", text).strip()


def load_diriginti(xlsx_path: Path) -> list[dict]:
    """Load diriginti de santier (site managers, PF only)."""
    rows = []
    if not xlsx_path.exists():
        print(f"[warn] {xlsx_path.name} not found — skipping diriginti")
        return rows

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            name = clean(ws.cell(r, 1).value)
            if not name or len(name) < 3:
                continue
            rows.append({
                "tip_autorizatie": "Diriginte de santier",
                "name": name,
                "nr_autorizatie": clean(ws.cell(r, 2).value),
                "data_emitere": clean(ws.cell(r, 3).value),
                "domenii": clean(ws.cell(r, 4).value),
                "email": clean(ws.cell(r, 7).value) if ws.cell(r, 7) else "",
                "telefon": clean(ws.cell(r, 8).value) if ws.cell(r, 8) else "",
            })
    except Exception as e:
        print(f"[error] Failed to parse diriginti: {e}", file=sys.stderr)
        return []

    print(f"[load] diriginti: {len(rows):,}")
    return rows


def load_rte(xlsx_path: Path) -> list[dict]:
    """Load RTE (technical execution managers, PF only)."""
    rows = []
    if not xlsx_path.exists():
        print(f"[warn] {xlsx_path.name} not found — skipping RTE")
        return rows

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            name = clean(ws.cell(r, 1).value)
            if not name or len(name) < 3:
                continue
            domenii = clean(ws.cell(r, 5).value)
            rows.append({
                "tip_autorizatie": "Responsabil tehnic cu executia",
                "name": name,
                "nr_autorizatie": clean(ws.cell(r, 2).value),
                "data_emitere": clean(ws.cell(r, 3).value),
                "valabilitate": clean(ws.cell(r, 4).value),
                "domenii": clean_html(domenii),
                "email": clean(ws.cell(r, 8).value) if ws.cell(r, 8) else "",
                "telefon": clean(ws.cell(r, 9).value) if ws.cell(r, 9) else "",
            })
    except Exception as e:
        print(f"[error] Failed to parse RTE: {e}", file=sys.stderr)
        return []

    print(f"[load] RTE: {len(rows):,}")
    return rows


def load_laboratoare(xlsx_path: Path) -> list[dict]:
    """Load laboratoare (testing labs, PJ with CUI)."""
    rows = []
    if not xlsx_path.exists():
        print(f"[warn] {xlsx_path.name} not found — skipping laboratoare")
        return rows

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            raw_name = clean(ws.cell(r, 1).value)
            if not raw_name or len(raw_name) < 3:
                continue
            cui = extract_cui(raw_name)
            rows.append({
                "tip_autorizatie": "Laborator autorizat ISC",
                "cui": cui,
                "name": clean_company_name(raw_name),
                "nr_autorizatie": clean(ws.cell(r, 3).value),
                "data_emitere": clean(ws.cell(r, 4).value),
                "domenii": clean(ws.cell(r, 5).value),
                "valabilitate": clean(ws.cell(r, 6).value) if ws.cell(r, 6) else "",
                "email": "",
                "telefon": "",
            })
    except Exception as e:
        print(f"[error] Failed to parse laboratoare: {e}", file=sys.stderr)
        return []

    print(f"[load] laboratoare: {len(rows):,}")
    return rows


def save_csv(rows: list[dict], output_path: Path, fieldnames: list) -> int:
    """Save rows to CSV."""
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"[save] {len(rows):,} rows -> {output_path.name}")
    return len(rows)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Scrape ISC registries from Excel files")
    parser.add_argument(
        "--consolidate",
        action="store_true",
        help="Export single consolidated CSV instead of per-category files",
    )
    args = parser.parse_args()

    print(f"[start] {datetime.now().isoformat()} — Scraping ISC registries")

    # Load all three registries
    diriginti = load_diriginti(XLSX_FILES["diriginti"])
    rte = load_rte(XLSX_FILES["rte"])
    laboratoare = load_laboratoare(XLSX_FILES["laboratoare"])

    total = len(diriginti) + len(rte) + len(laboratoare)
    print(f"[total] {total:,} records loaded (diriginti: {len(diriginti):,}, "
          f"RTE: {len(rte):,}, laboratoare: {len(laboratoare):,})")

    if args.consolidate:
        # Save as single consolidated file
        all_rows = diriginti + rte + laboratoare
        fieldnames = [
            "tip_autorizatie",
            "cui",
            "name",
            "nr_autorizatie",
            "data_emitere",
            "valabilitate",
            "domenii",
            "email",
            "telefon",
        ]
        save_csv(all_rows, OUTPUT_FILES["all"], fieldnames)
    else:
        # Save per-category files
        save_csv(
            diriginti,
            OUTPUT_FILES["diriginti"],
            ["tip_autorizatie", "name", "nr_autorizatie", "data_emitere", "domenii", "email", "telefon"],
        )
        save_csv(
            rte,
            OUTPUT_FILES["rte"],
            [
                "tip_autorizatie",
                "name",
                "nr_autorizatie",
                "data_emitere",
                "valabilitate",
                "domenii",
                "email",
                "telefon",
            ],
        )
        save_csv(
            laboratoare,
            OUTPUT_FILES["laboratoare"],
            ["tip_autorizatie", "cui", "name", "nr_autorizatie", "data_emitere", "valabilitate", "domenii", "email",
             "telefon"],
        )

    print(f"[done] {datetime.now().isoformat()}")


if __name__ == "__main__":
    main()
