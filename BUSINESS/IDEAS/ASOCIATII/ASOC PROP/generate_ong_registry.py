from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SOURCE_DATE = "2026-03-09"

FILE_CONFIG = {
    "09_03_2026Asociatii.xlsx": {
        "categorie": "asociatie",
        "source_url": "https://www.just.ro/wp-content/uploads/2026/03/09_03_2026Asociatii.xlsx",
    },
    "09_03_2026Fundatii.xlsx": {
        "categorie": "fundatie",
        "source_url": "https://www.just.ro/wp-content/uploads/2026/03/09_03_2026Fundatii.xlsx",
    },
    "09_03_2026Federatii.xlsx": {
        "categorie": "federatie",
        "source_url": "https://www.just.ro/wp-content/uploads/2026/03/09_03_2026Federatii.xlsx",
    },
    "09_03_2026Uniuni.xlsx": {
        "categorie": "uniune",
        "source_url": "https://www.just.ro/wp-content/uploads/2026/03/09_03_2026Uniuni.xlsx",
    },
    "09_03_2026PersJuridstr.xlsx": {
        "categorie": "persoana_juridica_straina",
        "source_url": "https://www.just.ro/wp-content/uploads/2026/03/09_03_2026PersJuridstr.xlsx",
    },
}

OUTPUT_FILE = Path("ONG_REGISTRU_NATIONAL.csv")
OUTPUT_ACTIVE_FILE = Path("ONG_ACTIVE.csv")
OUTPUT_SUMMARY_FILE = Path("ONG_SUMAR_JUDETE.csv")
OUTPUT_CITY_SUMMARY_FILE = Path("ONG_SUMAR_LOCALITATI.csv")
OUTPUT_COUNTY_DIR = Path("ong_pe_judete")

FIELDNAMES = [
    "categorie",
    "denumire",
    "numar_registru_national",
    "stare_actuala",
    "status_normalizat",
    "activ",
    "sediu_schimbat",
    "instanta_transfer",
    "tara",
    "judet",
    "localitate",
    "adresa",
    "scop_initial",
    "scop_modificari",
    "hg_utilitate_publica",
    "data_hg_utilitate_publica",
    "sursa_fisier",
    "sursa_url",
    "sursa_data_actualizare",
]


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        normalized = value.replace("_x000D_", " ")
        normalized = normalized.replace("\r", " ").replace("\n", " ")
        return " ".join(normalized.split())
    return str(value)


def normalize_status(value: str) -> str:
    cleaned = clean(value).lower()
    return cleaned or "activa"


def county_slug(value: str) -> str:
    cleaned = clean(value) or "NEPRECIZAT"
    ascii_value = unicodedata.normalize("NFKD", cleaned).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9]+", "_", ascii_value).strip("_").lower() or "neprecizat"


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, rows: list[dict[str, str]]) -> None:
    summary_rows = []
    grouped: dict[tuple[str, str], dict[str, int]] = {}

    for row in rows:
        key = (row["judet"] or "NEPRECIZAT", row["categorie"])
        stats = grouped.setdefault(key, {"total": 0, "active": 0, "inactive": 0})
        stats["total"] += 1
        if row["activ"] == "1":
            stats["active"] += 1
        else:
            stats["inactive"] += 1

    for (judet, categorie), stats in sorted(grouped.items()):
        summary_rows.append(
            {
                "judet": judet,
                "categorie": categorie,
                "total": stats["total"],
                "active": stats["active"],
                "inactive": stats["inactive"],
            }
        )

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["judet", "categorie", "total", "active", "inactive"])
        writer.writeheader()
        writer.writerows(summary_rows)


def write_city_summary(path: Path, rows: list[dict[str, str]]) -> None:
    grouped: dict[tuple[str, str], dict[str, int]] = {}

    for row in rows:
        key = (row["judet"] or "NEPRECIZAT", row["localitate"] or "NEPRECIZAT")
        stats = grouped.setdefault(key, {"total": 0, "asociatie": 0, "fundatie": 0, "federatie": 0, "uniune": 0, "persoana_juridica_straina": 0})
        stats["total"] += 1
        stats[row["categorie"]] += 1

    summary_rows = []
    for (judet, localitate), stats in sorted(grouped.items(), key=lambda item: (-item[1]["total"], item[0][0], item[0][1])):
        summary_rows.append(
            {
                "judet": judet,
                "localitate": localitate,
                "total_active": stats["total"],
                "asociatii_active": stats["asociatie"],
                "fundatii_active": stats["fundatie"],
                "federatii_active": stats["federatie"],
                "uniuni_active": stats["uniune"],
                "persoane_juridice_straine_active": stats["persoana_juridica_straina"],
            }
        )

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "judet",
                "localitate",
                "total_active",
                "asociatii_active",
                "fundatii_active",
                "federatii_active",
                "uniuni_active",
                "persoane_juridice_straine_active",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)


def write_county_files(rows: list[dict[str, str]]) -> None:
    OUTPUT_COUNTY_DIR.mkdir(exist_ok=True)
    county_groups: dict[str, list[dict[str, str]]] = {}

    for row in rows:
        county = row["judet"] or "NEPRECIZAT"
        county_groups.setdefault(county, []).append(row)

    for county, county_rows in county_groups.items():
        county_rows.sort(key=lambda row: (row["localitate"], row["denumire"], row["categorie"]))
        path = OUTPUT_COUNTY_DIR / f"ong_active_{county_slug(county)}.csv"
        write_csv(path, county_rows)


def load_rows(path: Path) -> Iterable[dict[str, str]]:
    config = FILE_CONFIG[path.name]
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]

    for row in rows:
        values = [clean(value) for value in row[: len(headers)]]
        if not any(values):
            continue

        data = dict(zip(headers, values))
        scop_parts = []
        for key in headers:
            if key.startswith("Modificari ale scopului") and data.get(key):
                scop_parts.append(data[key])

        yield {
            "categorie": config["categorie"],
            "denumire": data.get("Denumire", ""),
            "numar_registru_national": data.get("Numar inreg Reg National", ""),
            "stare_actuala": data.get("Starea actuala", ""),
            "status_normalizat": normalize_status(data.get("Starea actuala", "")),
            "activ": "1" if not clean(data.get("Starea actuala", "")) else "0",
            "sediu_schimbat": data.get("sediu schimbat", ""),
            "instanta_transfer": data.get("organizatia trece in circumscriptia altei instante", ""),
            "tara": data.get("Tara", ""),
            "judet": data.get("Judet", ""),
            "localitate": data.get("Localitate", ""),
            "adresa": data.get("Adresa", ""),
            "scop_initial": data.get("Scopul initial", ""),
            "scop_modificari": " | ".join(scop_parts),
            "hg_utilitate_publica": data.get("HG utilitate publica", ""),
            "data_hg_utilitate_publica": data.get("Data HG utilitate publica", ""),
            "sursa_fisier": path.name,
            "sursa_url": config["source_url"],
            "sursa_data_actualizare": SOURCE_DATE,
        }

    workbook.close()


def main() -> None:
    rows = []
    for file_name in FILE_CONFIG:
        path = Path(file_name)
        if not path.exists():
            raise FileNotFoundError(f"Lipseste fisierul: {file_name}")
        rows.extend(load_rows(path))

    rows.sort(key=lambda row: (row["judet"], row["localitate"], row["denumire"], row["categorie"]))
    write_csv(OUTPUT_FILE, rows)

    active_rows = [row for row in rows if row["activ"] == "1"]
    write_csv(OUTPUT_ACTIVE_FILE, active_rows)
    write_summary(OUTPUT_SUMMARY_FILE, rows)
    write_city_summary(OUTPUT_CITY_SUMMARY_FILE, active_rows)
    write_county_files(active_rows)

    print(f"Generated {OUTPUT_FILE}")
    print(f"Generated {OUTPUT_ACTIVE_FILE}")
    print(f"Generated {OUTPUT_SUMMARY_FILE}")
    print(f"Generated {OUTPUT_CITY_SUMMARY_FILE}")
    print(f"Generated county files in {OUTPUT_COUNTY_DIR}")


if __name__ == "__main__":
    main()