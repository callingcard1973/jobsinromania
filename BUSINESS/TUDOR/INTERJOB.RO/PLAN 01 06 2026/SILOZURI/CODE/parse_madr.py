#!/usr/bin/env python3
"""Parse MADR authorized storage facilities Excel files."""
import re
from openpyxl import load_workbook
from pathlib import Path
import csv

DATA_DIR = Path(__file__).parent / "DATA"
OUTPUT_CSV = Path(__file__).parent / "silos_master.csv"

def normalize_phone(phone_str):
    """Normalize RO phone to E.164 format."""
    if not phone_str:
        return ""
    phone = re.sub(r"[^0-9+]", "", str(phone_str).strip().upper())
    if not phone:
        return ""
    if phone.startswith("00"):
        phone = "+" + phone[2:]
    elif phone.startswith("0"):
        phone = "+40" + phone[1:]
    elif not phone.startswith("+"):
        phone = "+40" + phone
    return phone

def parse_excel_file(filepath):
    """Parse one county Excel file with multi-row facility structure."""
    facilities = []
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Skip header rows (usually first 8 rows)
        data_start = 8
        i = data_start

        while i < len(rows):
            row = rows[i]

            # Skip empty rows
            if not any(row):
                i += 1
                continue

            # Check if this is a facility row (has nr crt and auth number)
            if row[0] is not None and isinstance(row[0], (int, float)):
                # Facility block starts here
                # Row i: auth info + operator name + facility name + capacity
                # Row i+1: more operator/facility contact info
                # Row i+2: more contact info

                str(row[1] or "").strip() if len(row) > 1 else ""
                operator_name = str(row[2] or "").strip() if len(row) > 2 else ""
                facility_name = str(row[3] or "").strip() if len(row) > 3 else ""

                # Collect next 2 rows for contact info
                contact_info = ""
                if i + 1 < len(rows):
                    next_row = rows[i + 1]
                    contact_info += " ".join(str(v or "").strip() for v in next_row if v) + " "
                if i + 2 < len(rows):
                    next_row = rows[i + 2]
                    contact_info += " ".join(str(v or "").strip() for v in next_row if v)

                # Extract phone and email from contact info
                phones = re.findall(r"TEL\.?\s*([0-9+\-\s()]+)(?:\s|$)", contact_info.upper())
                emails = re.findall(r"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", contact_info)

                # Extract address (usually between name and phone)
                address_match = re.search(r"([A-Z][A-Za-z\s,0-9.]+)(?:TEL|$)", contact_info)
                address = address_match.group(1).strip() if address_match else ""

                phone = normalize_phone(phones[0]) if phones else ""
                email = emails[0].lower() if emails else ""

                # Use facility name if operator name not clear
                name = facility_name or operator_name

                if name:
                    facilities.append({
                        "county": filepath.stem,
                        "name": name,
                        "address": address,
                        "email": email,
                        "phone": phone,
                        "manager": operator_name if operator_name != facility_name else "",
                    })

                i += 3
            else:
                i += 1

    except Exception as e:
        print(f"ERROR parsing {filepath}: {e}")

    return facilities

# Process all county files
all_silos = []
for xlsx_file in sorted(DATA_DIR.glob("*.xlsx")):
    facilities = parse_excel_file(xlsx_file)
    all_silos.extend(facilities)
    print(f"[{xlsx_file.stem}] {len(facilities)} facilities")

print(f"\nTotal: {len(all_silos)}")

# Dedup by email+phone
deduped = {}
for silo in all_silos:
    key = (silo["email"], silo["phone"])
    if key not in deduped:
        deduped[key] = silo

print(f"After dedup: {len(deduped)}")

# Filter: must have email or phone
with_contact = [s for s in deduped.values() if s["email"] or s["phone"]]
print(f"With email/phone: {len(with_contact)}")

# Write to CSV
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["county", "name", "address", "email", "phone", "manager"])
    writer.writeheader()
    for silo in sorted(with_contact, key=lambda x: (x["county"], x["name"])):
        writer.writerow(silo)

print(f"\nOK: {OUTPUT_CSV}")
print("Sample:")
for silo in sorted(with_contact, key=lambda x: x["county"])[:5]:
    print(f"  {silo['county']}: {silo['name']}")
    if silo['email']:
        print(f"    Email: {silo['email']}")
    if silo['phone']:
        print(f"    Phone: {silo['phone']}")
