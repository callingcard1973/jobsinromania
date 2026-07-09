#!/usr/bin/env python3
"""Parse MADR Excel files CORRECTLY - multi-row facility blocks."""
import re
from openpyxl import load_workbook
from pathlib import Path
import csv

DATA_DIR = Path(__file__).parent.parent / "DATA"
OUTPUT_CSV = Path(__file__).parent.parent / "silos_master_correct.csv"

def extract_phone(text):
    """Extract phone from text like 'TEL.0742123456' or '0742 123 456'."""
    if not text:
        return ""
    # Look for phone patterns
    phones = re.findall(r'(?:TEL\.?\s*)?(\d{3,}(?:[\s\-\(\)]+\d+)*)', str(text).upper())
    if phones:
        # Clean: remove spaces, parens, dashes
        phone = re.sub(r'[\s\-\(\)]', '', phones[0])
        # Normalize to +40
        if phone.startswith('0'):
            phone = '+40' + phone[1:]
        elif not phone.startswith('+'):
            phone = '+40' + phone
        return phone
    return ""

def extract_email(text):
    """Extract email from text."""
    if not text:
        return ""
    emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', str(text))
    return emails[0].lower() if emails else ""

def parse_xlsx(filepath):
    """Parse one Excel file - CORRECT multi-row handling."""
    facilities = []
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Find header rows (usually row 3-4 contain column names)
        header_row = None
        for i, row in enumerate(rows[:10]):
            row_text = ' '.join(str(c or '') for c in row).lower()
            if 'autorizatia' in row_text or 'denumire' in row_text or 'nr.crt' in row_text:
                header_row = i
                break

        if header_row is None:
            header_row = 3  # Default

        # Data starts after headers + blank rows
        data_start = header_row + 5
        i = data_start

        while i < len(rows):
            row = rows[i]
            # Skip empty rows
            if not any(row):
                i += 1
                continue

            # Check if this is a facility data row (has row number or auth code)
            str(row[0] or '').strip()
            col1 = str(row[1] or '').strip()

            # Auth number usually in col1 (format: XX00123)
            if col1 and re.match(r'^[A-Z]{2}\d{5}', col1):
                # This is a facility block start
                # Row i: auth + operator name + facility name + capacity
                auth_num = col1
                operator_name = str(row[2] or '').strip()
                facility_name = str(row[3] or '').strip()
                capacity = str(row[4] or '').strip() if len(row) > 4 else ''

                # Next 2 rows have contact info
                contact_lines = []
                if i + 1 < len(rows):
                    contact_lines.append(' '.join(str(c or '') for c in rows[i + 1]))
                if i + 2 < len(rows):
                    contact_lines.append(' '.join(str(c or '') for c in rows[i + 2]))

                contact_text = ' '.join(contact_lines)

                # Extract address (text before TEL or EMAIL)
                address_match = re.search(r'^(.+?)(?:TEL|EMAIL|$)', contact_text, re.IGNORECASE)
                address = address_match.group(1).strip() if address_match else ''
                address = re.sub(r'\s+', ' ', address)  # Clean whitespace

                # Extract phone and email
                phone = extract_phone(contact_text)
                email = extract_email(contact_text)

                # Use facility name if available, else operator
                name = facility_name or operator_name
                if name and name.strip():
                    facilities.append({
                        'county': filepath.stem,
                        'name': name.strip(),
                        'address': address[:100],  # Trim long addresses
                        'email': email,
                        'phone': phone,
                        'manager': operator_name.strip() if operator_name != facility_name else '',
                        'auth_num': auth_num,
                        'capacity_tonnes': capacity,
                    })

                i += 3  # Skip the contact rows
            else:
                i += 1

    except Exception as e:
        print(f"ERROR {filepath.stem}: {e}")

    return facilities

def main():
    print("Parsing MADR Excel files CORRECTLY...\n")

    all_facilities = []
    for xlsx_file in sorted(DATA_DIR.glob("*.xlsx")):
        facilities = parse_xlsx(xlsx_file)
        all_facilities.extend(facilities)
        if facilities:
            print(f"[{xlsx_file.stem}] {len(facilities)} facilities")

    print(f"\nTotal: {len(all_facilities)}")

    # Dedup by (name, phone, email)
    seen = {}
    for fac in all_facilities:
        key = (fac['name'].lower(), fac['phone'], fac['email'])
        if key not in seen:
            seen[key] = fac

    print(f"After dedup: {len(seen)}")

    # Stats
    with_email = len([f for f in seen.values() if f['email']])
    with_phone = len([f for f in seen.values() if f['phone']])
    with_both = len([f for f in seen.values() if f['email'] and f['phone']])

    print("\nContact coverage:")
    print(f"  Email: {with_email} ({with_email*100/len(seen):.1f}%)")
    print(f"  Phone: {with_phone} ({with_phone*100/len(seen):.1f}%)")
    print(f"  Both: {with_both} ({with_both*100/len(seen):.1f}%)")

    # Write CSV
    fieldnames = ['county', 'name', 'address', 'email', 'phone', 'manager', 'auth_num', 'capacity_tonnes']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for fac in sorted(seen.values(), key=lambda x: (x['county'], x['name'])):
            writer.writerow(fac)

    print(f"\nOK: {OUTPUT_CSV}")
    print("Sample:")
    for fac in list(seen.values())[:5]:
        print(f"  {fac['name']}")
        if fac['email']:
            print(f"    EMAIL: {fac['email']}")
        if fac['phone']:
            print(f"    PHONE: {fac['phone']}")

if __name__ == "__main__":
    main()
