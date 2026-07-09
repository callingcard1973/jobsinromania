#!/usr/bin/env python3
"""Extract all authorized storage facilities from MADR county files."""
import requests
from openpyxl import load_workbook
from pathlib import Path
import csv
from io import BytesIO
import time

BASE_URL = "https://www.madr.ro"
OUTPUT_DIR = Path(__file__).parent / "DATA"
OUTPUT_CSV = OUTPUT_DIR / "silos_master.csv"

COUNTIES = [
    "Alba", "Arad", "Arges", "Bacau", "Bihor", "Bistrita-Nasaud", "Botosani",
    "Braila", "Brasov", "Bucuresti", "Buzau", "Calarasi", "Caras-Severin",
    "Cluj", "Constanta", "Covasna", "Dambovita", "Dolj", "Galati", "Giurgiu",
    "Gorj", "Harghita", "Hunedoara", "Ialomita", "Iasi", "Ifov",
    "Ilfov", "Maramures", "Mehedinti", "Mures", "Neamt", "Olt",
    "Prahova", "Salaj", "Satu-Mare", "Sibiu", "Suceava", "Teleorman",
    "Timis", "Tulcea", "Valcea", "Vaslui", "Vrancea"
]

def download_and_parse_county(county_name):
    """Download and parse a single county Excel file."""
    url_patterns = [
        f"{BASE_URL}/docs/depozite/2025/{county_name}-2025.xlsx",
        f"{BASE_URL}/docs/depozite/2025/{county_name}-2025.xls",
    ]

    data = []
    for url in url_patterns:
        try:
            print(f"  {url}")
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                try:
                    wb = load_workbook(BytesIO(resp.content), data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))

                    if len(rows) < 2:
                        print("    [no data]")
                        continue

                    headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

                    for row in rows[1:]:
                        if not any(row):
                            continue

                        row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}

                        facility = {
                            "county": county_name,
                            "name": str(row_dict.get("denumire", "") or row_dict.get("name", "") or "").strip(),
                            "address": str(row_dict.get("adresa", "") or row_dict.get("address", "") or "").strip(),
                            "email": str(row_dict.get("email", "") or "").strip().lower() if row_dict.get("email") else "",
                            "phone": str(row_dict.get("telefon", "") or row_dict.get("phone", "") or "").strip() if row_dict.get("telefon") else "",
                            "manager": str(row_dict.get("manager", "") or row_dict.get("manager_name", "") or "").strip(),
                        }
                        if facility["name"]:
                            data.append(facility)

                    print(f"    [parsed {len(data)} facilities]")
                    return data
                except Exception as e:
                    print(f"    [parse error: {e}]")
                    continue
        except requests.exceptions.RequestException:
            print("    [connection error]")
            continue

    print("  [not found]")
    return []

def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    all_silos = []

    print("Downloading MADR county files...\n")
    for county in COUNTIES:
        try:
            silos = download_and_parse_county(county)
            all_silos.extend(silos)
            time.sleep(0.3)
        except Exception as e:
            print(f"  [{county} error: {e}]")

    print(f"\n[total extracted: {len(all_silos)} before dedup]")

    deduped = {}
    for silo in all_silos:
        key = (silo["name"].lower(), silo["email"], silo["phone"])
        if key not in deduped:
            deduped[key] = silo

    print(f"[after dedup: {len(deduped)} unique]")

    OUTPUT_DIR.mkdir(exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["county", "name", "address", "email", "phone", "manager"])
        writer.writeheader()
        for silo in sorted(deduped.values(), key=lambda x: x["county"]):
            writer.writerow(silo)

    print(f"\n[OK] {OUTPUT_CSV}")
    print(f"     Rows: {len(deduped)}")

    print("\n[Sample]")
    for i, silo in enumerate(list(deduped.values())[:5], 1):
        print(f"  {i}. {silo['name']} | {silo['county']}")
        if silo['email']:
            print(f"     {silo['email']}")
        if silo['phone']:
            print(f"     {silo['phone']}")

if __name__ == "__main__":
    main()
