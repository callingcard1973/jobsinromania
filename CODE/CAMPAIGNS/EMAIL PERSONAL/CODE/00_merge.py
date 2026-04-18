import csv
import os
import re

# Google Contacts format (standard)
SOURCES = [
    '../DATA/contacts-fruitnature.csv',
    '../DATA/FRUINATURE2/contacts_FRuitnature2.csv',
    '../DATA/FRUINATURE3/contacts_fruitnature3.csv',
    '../DATA/FRUINATURE4/contacts_fruitnature4.csv',
]

# Yahoo Contacts format (Company, Job Title, Email, Mobile...)
YAHOO_SOURCES = [
    '../DATA/APAMINERALA YAHOO/contacts_apaminerala_yahoo.csv',
]

# Plain email lists (one email per line, no header)
EMAIL_ONLY_SOURCES = [
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/6 ian 2015/Desktop/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi1.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/6 ian 2015/Desktop/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi2.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/6 ian 2015/Desktop/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi5.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/6 ian 2015/Desktop/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi6.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi1.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri de anul nou/practicieni insolventa/Doar emailuri/insolventi2.csv', 'insolventa_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/6 ian 2015/Desktop/Emailuri de anul nou/cadastristi/cadastristi ancpi.csv', 'cadastristi_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri de anul nou/cadastristi/cadastristi ancpi.csv', 'cadastristi_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 21 martie 2016/Sectiunea- 5 - SAI - Societati de administrare a investitiilor --Inapoi - Sheet1 (1) - numai emailuri.csv', 'investitii_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/Desktop 18 iulie 2016/societati investitii romania doar emailuri.csv', 'investitii_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/Norway/emails norway.csv', 'norway'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/sweden/Flowers/emails sweden.csv', 'sweden'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/AUSTRIA/AUSTRIA DISTRIBUITORI/AUSTRIA WHOLESALERS FRUIT VEGETABLES txt 1.csv', 'wholesale_at'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/fruits_et_legumes rungis France juste emails - column.txt', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/RUNGIS1.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/SEMMARIS RUNGIS.csv', 'wholesale_fr'),
    # plain email scan — Austria, Romania, Poland, Russia
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/AUSTRIA/AUSTRIA DISTRIBUITORI/kontaktdatenverzeichnis csv 03 07 2023.csv', 'wholesale_at'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/Emailuri agricultura/austria anunturi agenti comerciali.csv', 'wholesale_at'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/Emailuri agricultura/infoferma total emailuri.csv', 'agricultura_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/Emailuri agricultura/exporters importers spices from polish website.csv', 'wholesale_pl'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/Emailuri agricultura/Rusia Fruit Inform/companii rusesti si rusofone.txt', 'wholesale_ru'),
]

# "Email","URL" format (two columns)
EMAIL_URL_SOURCES = [
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/Emailuri agricultura/Romania/produse apicole/apicole.csv', 'agricultura_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri agricultura/Romania/produse apicole/apicole.csv', 'agricultura_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri agricultura/Romania/Plante medicinale/medicinale.csv', 'agricultura_ro'),
    ('E:/BACKUPS/backup hdd extern vechi 27 nov 2018/Desktopuri/desktop 4 decembrie HP prioritatea 2/Emailuri agricultura/Romania/produse traditionale/prod trad.csv', 'agricultura_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/Emailuri agricultura/Romania/legume-fructe/legume fructe.csv', 'legume_fructe_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/Emailuri agricultura/Romania/produse apicole/apicole.csv', 'agricultura_ro'),
    ('I:/BUSINESS/ALUMINUM/EMAIL/centre colectare deseuri in total ro.csv', 'reciclare_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/ROMANIA/Emailuri Unitati Cazare Romania - Sheet1 (1).csv', 'horeca_ro'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/AUSTRIA/AUSTRIA DISTRIBUITORI/AUSTRIA WHOLESALERS FRUIT VEGETABLES.csv', 'wholesale_at'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/BELGIUM/FEBEV is the National Belgian Federation of slaughterhouses, cutting plants and wholesalers for pork, bovine, sheep and equidae..csv', 'wholesale_be'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/Emailuri agricultura/emailuri wholesale germania berlin fructe.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FINLAND TEURASTAMO WHOLESALE.csv', 'wholesale_fi'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FIRENZE WHOLESALERS.csv', 'wholesale_it'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/GENOVA WHOLESALERS.csv', 'wholesale_it'),
    # Belgium
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/BELGIUM/MABRU BRUXELLES.csv', 'wholesale_be'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/BELGIUM/BELGIUM 2.csv', 'wholesale_be'),
    # Denmark
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/DENMARK/DENMARK CONTACT.csv', 'wholesale_dk'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/DENMARK/ORGANIC DENMARK.csv', 'wholesale_dk'),
    # UK / Scotland
    ('E:/BACKUPS/Desktop/DESKTOP VECHI/AGROEVOLUTION/BUYERS/EUROPE/Bagate deja in sendinblue/uk fruit wholesale.csv', 'wholesale_uk'),
    ('E:/BACKUPS/Desktop/DESKTOP VECHI/AGROEVOLUTION/BUYERS/EUROPE/Bagate deja in sendinblue/scottish wholesale association members.csv', 'wholesale_uk'),
    ('E:/BACKUPS/Desktop/DESKTOP VECHI/AGROEVOLUTION/BUYERS/EUROPE/Bagate deja in sendinblue/vienna wholesale association members.csv', 'wholesale_at'),
    # Germany — Grossmarkt per city (15 markets)
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/BERLIN GROSSMARKT .csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/BOLZANO GROSSMARKT.csv', 'wholesale_it'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/DUISBURG GROSSMARKT.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/ESSEN.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/FRANKFURT.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/FRUCHTHOF BERLIN.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/GROSSMARKT BREMEN.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/GROSSMARKT DORTMUND.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/HAMBURG.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/HANNOVER.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/KOLN.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/LEIPZIG.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/MUNCHEN.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/NECUNOSCUTI.csv', 'wholesale_de'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/GERMANY/EMAIL GRABBER/STUTTGART.csv', 'wholesale_de'),
    # France — MIN cities (14 markets)
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/1.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/AGEN.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/ANGERS.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/ANJOU.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/AVIGNON.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/Bordeaux-Brienne.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/CAEN.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/GRENOBLE ENTREPRISES.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/Grenoble 2.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/LILLE.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/MIN BOrdeaux Brienne.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/Nice Cote d;Azur.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/PERPIGNAN ACHETEURS.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/PERPIGNAN GROSSISTES.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/PRODUCTEURS/Perpignan Producteurs.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/STRASBOURG.csv', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/TOURS.csv', 'wholesale_fr'),
]

# XLSX sources (Vienna wholesale, RNCA cooperatives)
XLSX_SOURCES = [
    {
        'path': 'I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/Contact data wholesale markets/vienna contact database kontaktdatenverzeichnis.xls.xlsx',
        'sheet': None, 'skip_rows': 1,
        'col_map': {'E-Mail': 'E-Mail', 'Firma': 'Firma', 'Telefon': 'Telefon', 'Homepage': 'Homepage'},
        'label': 'wholesale_at_vienna',
    },
]

# PDF sources (email regex scan over full text)
PDF_SOURCES = [
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/PROSPECTING/FINLAND/FRANCE/annuaire2016-fm rungis (2023_02_08 05_51_45 UTC).pdf', 'wholesale_fr'),
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/France/fruits_et_legumes grossites Rungis.pdf', 'wholesale_fr'),
]

# Binary DOC sources (old Word format, binary email scan)
DOC_SOURCES = [
    ('I:/BUSINESS/OIPA EXPORT 2023/ALL/HAMBARUL ROMANESC/Legume fructe agri zootehnie 2017/France/emails grossistes fruits et legumes rungis.doc', 'wholesale_fr'),
]

OUTPUT = '../DATA/merged.csv'

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

def make_row(email='', org='', phone='', website='', notes='', label=''):
    return {
        'First Name': '', 'Middle Name': '', 'Last Name': '',
        'Organization Name': org, 'Organization Title': '',
        'Notes': notes,
        'E-mail 1 - Value': email.strip().lower(),
        'E-mail 2 - Value': '', 'Phone 1 - Value': phone,
        'Website 1 - Value': website, 'Labels': label or '* myContacts',
        'Birthday': '',
    }

def convert_yahoo_row(row):
    email = row.get('Email', '').strip() or row.get('Work Email', '').strip() or row.get('Home Email', '').strip()
    phone = row.get('Mobile', '').strip() or row.get('Phone', '').strip() or row.get('Work', '').strip() or row.get('Home', '').strip()
    return {
        'First Name': row.get('First Name', ''),
        'Middle Name': row.get('Middle Name', ''),
        'Last Name': row.get('Last Name', ''),
        'Organization Name': row.get('Company', ''),
        'Organization Title': row.get('Job Title', ''),
        'Notes': row.get('Notes', ''),
        'E-mail 1 - Value': email.strip().lower(),
        'E-mail 2 - Value': row.get('Home Email', '').strip().lower() if email.lower() != row.get('Home Email', '').strip().lower() else '',
        'Phone 1 - Value': phone,
        'Website 1 - Value': row.get('website', ''),
        'Labels': '* myContacts',
        'Birthday': row.get('Birthday', ''),
    }

def load_email_only(path, label):
    rows = []
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.strip()
            m = EMAIL_RE.search(line)
            if m:
                rows.append(make_row(email=m.group(), label=label))
    return rows

def load_email_url(path, label):
    rows = []
    with open(path, encoding='utf-8', errors='replace') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames and any('mail' in (fn or '').lower() or 'Email' in (fn or '') for fn in reader.fieldnames):
            for row in reader:
                email = row.get('Email', row.get('email', '')).strip().lower()
                url = row.get('URL', row.get('url', row.get('Website', ''))).strip()
                if email:
                    rows.append(make_row(email=email, website=url, label=label))
        else:
            # fallback: plain email scan
            f.seek(0)
            for line in f:
                m = EMAIL_RE.search(line)
                if m:
                    rows.append(make_row(email=m.group(), label=label))
    return rows

def load_xlsx(path, sheet, skip_rows, col_map, label):
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl needed for XLSX")
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active if not sheet else (wb[sheet] if sheet in wb.sheetnames else wb.active)
    rows_out = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        if headers is None:
            headers = [str(v) if v else '' for v in row]
            continue
        d = dict(zip(headers, row))
        email = str(d.get(col_map.get('E-Mail', 'E-Mail'), '') or '').strip().lower()
        if not EMAIL_RE.search(email):
            continue
        rows_out.append(make_row(
            email=email,
            org=str(d.get(col_map.get('Firma', ''), '') or ''),
            phone=str(d.get(col_map.get('Telefon', ''), '') or ''),
            website=str(d.get(col_map.get('Homepage', ''), '') or ''),
            label=label,
        ))
    wb.close()
    return rows_out

def load_pdf_emails(path, label):
    try:
        import pdfplumber
    except ImportError:
        print("pip install pdfplumber needed for PDF")
        return []
    rows = []
    seen = set()
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for m in EMAIL_RE.findall(text):
                e = m.lower()
                if e not in seen:
                    seen.add(e)
                    rows.append(make_row(email=e, label=label))
    return rows

def load_doc_emails(path, label):
    rows = []
    seen = set()
    with open(path, 'rb') as f:
        content = f.read().decode('latin-1', errors='replace')
    for m in EMAIL_RE.findall(content):
        e = m.lower()
        if e not in seen:
            seen.add(e)
            rows.append(make_row(email=e, label=label))
    return rows

def ensure_fields(all_fields_ordered, seen_fields, row_keys):
    for field in row_keys:
        if field not in seen_fields:
            all_fields_ordered.append(field)
            seen_fields.add(field)

def main():
    all_rows = []
    all_fields_ordered = []
    seen_fields = set()

    # Google Contacts
    for path in SOURCES:
        if not os.path.exists(path):
            print(f"SKIP (not found): {path}")
            continue
        with open(path, encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            ensure_fields(all_fields_ordered, seen_fields, reader.fieldnames)
            rows = list(reader)
            all_rows.extend(rows)
        print(f"Google  {len(rows):>6} | {os.path.basename(path)}")

    # Yahoo Contacts
    for path in YAHOO_SOURCES:
        if not os.path.exists(path):
            print(f"SKIP (not found): {path}")
            continue
        with open(path, encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            rows = [convert_yahoo_row(r) for r in reader]
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"Yahoo   {len(rows):>6} | {os.path.basename(path)}")

    # Plain email lists
    seen_paths = set()
    for path, label in EMAIL_ONLY_SOURCES:
        if path in seen_paths or not os.path.exists(path):
            continue
        seen_paths.add(path)
        rows = load_email_only(path, label)
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"Email   {len(rows):>6} | {os.path.basename(path)} [{label}]")

    # Email+URL format
    seen_paths2 = set()
    for path, label in EMAIL_URL_SOURCES:
        if path in seen_paths2 or not os.path.exists(path):
            continue
        seen_paths2.add(path)
        rows = load_email_url(path, label)
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"EU      {len(rows):>6} | {os.path.basename(path)} [{label}]")

    # XLSX sources
    for src in XLSX_SOURCES:
        if not os.path.exists(src['path']):
            print(f"SKIP (not found): {os.path.basename(src['path'])}")
            continue
        rows = load_xlsx(src['path'], src['sheet'], src['skip_rows'], src['col_map'], src['label'])
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"XLSX  {len(rows):>6} | {os.path.basename(src['path'])} [{src['label']}]")

    # PDF sources
    seen_paths_pdf = set()
    for path, label in PDF_SOURCES:
        if path in seen_paths_pdf or not os.path.exists(path):
            seen_paths_pdf.add(path)
            continue
        seen_paths_pdf.add(path)
        rows = load_pdf_emails(path, label)
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"PDF   {len(rows):>6} | {os.path.basename(path)} [{label}]")

    # Binary DOC sources
    seen_paths_doc = set()
    for path, label in DOC_SOURCES:
        if path in seen_paths_doc or not os.path.exists(path):
            seen_paths_doc.add(path)
            continue
        seen_paths_doc.add(path)
        rows = load_doc_emails(path, label)
        if rows:
            ensure_fields(all_fields_ordered, seen_fields, rows[0].keys())
        all_rows.extend(rows)
        print(f"DOC   {len(rows):>6} | {os.path.basename(path)} [{label}]")

    with open(OUTPUT, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=all_fields_ordered, extrasaction='ignore')
        writer.writeheader()
        for row in all_rows:
            out = {k: '' for k in all_fields_ordered}
            out.update({k: v for k, v in row.items() if k in seen_fields})
            writer.writerow(out)

    print(f"\nTotal merged: {len(all_rows)} rows -> {OUTPUT}")

if __name__ == '__main__':
    main()
