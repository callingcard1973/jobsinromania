#!/usr/bin/env python3
"""
XLSX Converter - Convert CSV to Excel with formatting
Usage: python3 xlsx_converter.py <input.csv> [--output file.xlsx] [--sheet name]
"""
import sys
sys.path.insert(0, '/opt/ACTIVE/SCRAPERS/EUROPE/SCRIPTS/SHARED')
from skills_common import to_ascii, sanitize, FIELD_LIMITS

import os
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Try openpyxl, fall back to basic xlsx creation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================
# CSV READING
# ============================================================

def read_csv(filepath: str) -> Tuple[List[str], List[List]]:
    """Read CSV file and return headers and rows"""
    headers = []
    rows = []

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)

    return headers, rows

def detect_column_types(headers: List[str], rows: List[List]) -> Dict[int, str]:
    """Detect column types for formatting"""
    types = {}

    for i, header in enumerate(headers):
        header_lower = header.lower()

        # Check header name
        if 'email' in header_lower:
            types[i] = 'email'
        elif 'phone' in header_lower or 'tel' in header_lower:
            types[i] = 'phone'
        elif 'date' in header_lower or 'time' in header_lower:
            types[i] = 'datetime'
        elif 'url' in header_lower or 'link' in header_lower:
            types[i] = 'url'
        elif 'price' in header_lower or 'cost' in header_lower or 'amount' in header_lower:
            types[i] = 'currency'
        elif 'count' in header_lower or 'num' in header_lower or 'qty' in header_lower:
            types[i] = 'number'
        elif 'score' in header_lower or 'rate' in header_lower or 'percent' in header_lower:
            types[i] = 'percentage'
        else:
            types[i] = 'text'

    return types

# ============================================================
# XLSX CREATION WITH OPENPYXL
# ============================================================

def create_xlsx_openpyxl(
    headers: List[str],
    rows: List[List],
    output_path: str,
    sheet_name: str = "Data",
    add_chart: bool = False,
) -> str:
    """Create XLSX with openpyxl"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Detect column types
    col_types = detect_column_types(headers, rows)

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = thin_border

            # Apply formatting based on type
            col_type = col_types.get(col_idx, 'text')
            if col_type == 'number':
                try:
                    cell.value = int(value)
                except Exception:
                    pass
            elif col_type == 'currency':
                try:
                    cell.value = float(value.replace(',', '').replace('$', '').replace('€', ''))
                    cell.number_format = '#,##0.00'
                except Exception:
                    pass
            elif col_type == 'percentage':
                try:
                    val = float(value.replace('%', ''))
                    cell.value = val / 100
                    cell.number_format = '0.0%'
                except Exception:
                    pass

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col - 1]))
        for row in range(2, min(len(rows) + 2, 100)):  # Check first 100 rows
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    ws_summary['A3'] = "Total Rows:"
    ws_summary['B3'] = len(rows)
    ws_summary['A4'] = "Total Columns:"
    ws_summary['B4'] = len(headers)
    ws_summary['A5'] = "Generated:"
    ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Column summary
    ws_summary['A7'] = "Columns:"
    ws_summary['A7'].font = Font(bold=True)
    for i, header in enumerate(headers, 8):
        ws_summary[f'A{i}'] = header
        ws_summary[f'B{i}'] = col_types.get(i - 8, 'text')

    wb.save(output_path)
    return output_path

# ============================================================
# BASIC XLSX CREATION (NO DEPENDENCIES)
# ============================================================

def create_xlsx_basic(
    headers: List[str],
    rows: List[List],
    output_path: str,
    sheet_name: str = "Data",
) -> str:
    """Create basic XLSX without openpyxl (XML-based)"""
    import zipfile
    from xml.etree.ElementTree import Element, SubElement, tostring

    # Create workbook structure
    workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheets><sheet name="{sheet}" sheetId="1" r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></sheets>
</workbook>'''.format(sheet=sheet_name)

    # Create sheet data
    def col_letter(n):
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    sheet_data = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                  '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
                  '<sheetData>']

    # Header row
    sheet_data.append('<row r="1">')
    for i, header in enumerate(headers, 1):
        cell_ref = f"{col_letter(i)}1"
        escaped = str(header).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        sheet_data.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{escaped}</t></is></c>')
    sheet_data.append('</row>')

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        sheet_data.append(f'<row r="{row_idx}">')
        for col_idx, value in enumerate(row, 1):
            cell_ref = f"{col_letter(col_idx)}{row_idx}"
            escaped = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            sheet_data.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{escaped}</t></is></c>')
        sheet_data.append('</row>')

    sheet_data.append('</sheetData></worksheet>')
    sheet_xml = '\n'.join(sheet_data)

    # Content types
    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>'''

    # Relationships
    rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''

    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>'''

    # Create ZIP (XLSX)
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', rels)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)

    return output_path

# ============================================================
# MULTI-FILE CONVERSION
# ============================================================

def convert_multiple(
    input_files: List[str],
    output_path: str,
) -> str:
    """Convert multiple CSVs to single XLSX with multiple sheets"""
    if not HAS_OPENPYXL:
        print("Error: openpyxl required for multi-sheet conversion")
        return None

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for filepath in input_files:
        if not os.path.exists(filepath):
            continue

        sheet_name = Path(filepath).stem[:31]  # Excel limit
        headers, rows = read_csv(filepath)

        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    return output_path

# ============================================================
# MAIN
# ============================================================

def main():
    args = sys.argv[1:]

    if not args or '-h' in args or '--help' in args:
        print(f"""
{'='*60}
XLSX CONVERTER
{'='*60}

Usage: xlsx_converter.py <input.csv> [options]
       xlsx_converter.py file1.csv file2.csv --output combined.xlsx

Options:
  --output FILE     Output file path (default: input_name.xlsx)
  --sheet NAME      Sheet name (default: Data)
  --basic           Use basic mode (no openpyxl required)

Features (with openpyxl):
  - Auto-column width
  - Header formatting
  - Auto-filter
  - Freeze panes
  - Summary sheet
  - Type detection (email, phone, currency, etc.)

Status: openpyxl {'available' if HAS_OPENPYXL else 'NOT INSTALLED'}

Examples:
  xlsx_converter.py contacts.csv
  xlsx_converter.py data.csv --output report.xlsx --sheet "Q4 Data"
  xlsx_converter.py file1.csv file2.csv file3.csv --output combined.xlsx
""")
        return

    # Parse arguments
    input_files = []
    output_path = None
    sheet_name = "Data"
    use_basic = '--basic' in args or not HAS_OPENPYXL

    for i, arg in enumerate(args):
        if arg == '--output' and i + 1 < len(args):
            output_path = args[i + 1]
        elif arg == '--sheet' and i + 1 < len(args):
            sheet_name = args[i + 1]
        elif arg.endswith('.csv') and os.path.exists(arg):
            input_files.append(arg)

    if not input_files:
        print("Error: No valid CSV files found")
        return

    # Default output path
    if not output_path:
        output_path = Path(input_files[0]).stem + '.xlsx'

    print(f"\n{'='*60}")
    print(f"XLSX CONVERTER")
    print(f"{'='*60}\n")
    print(f"Input files: {len(input_files)}")
    print(f"Output: {output_path}")
    print(f"Mode: {'basic' if use_basic else 'openpyxl'}")

    if len(input_files) == 1:
        # Single file conversion
        headers, rows = read_csv(input_files[0])
        print(f"Rows: {len(rows)}, Columns: {len(headers)}")

        if use_basic:
            create_xlsx_basic(headers, rows, output_path, sheet_name)
        else:
            create_xlsx_openpyxl(headers, rows, output_path, sheet_name)
    else:
        # Multi-file conversion
        if use_basic:
            print("Error: Multi-file conversion requires openpyxl")
            return

        convert_multiple(input_files, output_path)
        print(f"Sheets created: {len(input_files)}")

    print(f"\nSaved to: {output_path}")
    print(f"Size: {os.path.getsize(output_path) / 1024:.1f} KB")
    print(f"\n{'='*60}\n")

if __name__ == '__main__':
    main()
